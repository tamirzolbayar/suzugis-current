from __future__ import annotations

import json
import math
import sqlite3
import struct
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
GPKG_PATH = ROOT / "data" / "qgis_real_roads" / "shoin_real_roads.gpkg"
QGIS_EXCEL_PATH = ROOT / "data" / "excel" / "qgis_construction_list.xlsx"
BASE_GEOJSON_PATH = ROOT / "data" / "geojson" / "suzu_sample.geojson"
BASE_EXCEL_PATH = ROOT / "data" / "excel" / "restriction_list.xlsx"
OUT_GEOJSON_PATH = ROOT / "data" / "geojson" / "suzu_with_real_roads.geojson"
OUT_EXCEL_PATH = ROOT / "data" / "excel" / "restriction_list_with_real_roads.xlsx"

KEY_ID = "\u898f\u5236ID"
KEY_DISTRICT = "\u91cd\u70b9\u5730\u533a"
KEY_SOURCE = "\u5143\u30c7\u30fc\u30bf"
KEY_EXTRACT = "\u62bd\u51fa\u6761\u4ef6"

COL_TITLE = "\u5de5\u4e8b\u540d"
COL_WORK_TYPE = "\u5de5\u4e8b\u7a2e\u5225"
COL_RESTRICTION_TYPE = "\u898f\u5236\u7a2e\u5225"
COL_START = "\u958b\u59cb\u65e5"
COL_END = "\u7d42\u4e86\u65e5"
COL_CONTRACTOR = "\u65bd\u5de5\u8005"
COL_PROGRESS = "\u9032\u6357\u7387"
COL_NOTE = "\u5099\u8003"

COL_ITEM_STATUS = "\u9805\u76ee\u72b6\u614b"
APP_REAL_ID = "\u5b9fID"
SRC_ROAD_NAME = "\u9053\u8def\u540d\u79f0"
SRC_ASSESSMENT = "\u67fb\u5b9a\u756a\u53f7"
SRC_DISTRICT = "\u5730\u533a"
SRC_WORK = "\u5de5\u4e8b\u533a\u5206"
SRC_ID = "ID"
SRC_SEGMENT = "\u533a\u5206"

EXCEL_ITEM_NAME = "\u500b\u5225\u9805\u76ee\u540d"
EXCEL_ROAD_NAME = "\u500b\u5225\u9053\u8def\u540d"
EXCEL_BRIDGE_NAME = "\u6a4b\u6881\u540d"
EXCEL_DISTRICT = "\u5730\u533a"
EXCEL_WORK = "\u5de5\u4e8b\u533a\u5206"
EXCEL_ASSESSMENT = "\u67fb\u5b9a\u756a\u53f7"
EXCEL_PLACE = "\u7b87\u6240\u540d"
EXCEL_LENGTH = "\u5fa9\u65e7\u5ef6\u9577_m"
EXCEL_WIDTH = "\u5e45\u54e1_m"
EXCEL_CONTENT = "\u5de5\u4e8b\u5185\u5bb9"
EXCEL_NUMBER_TYPE = "\u756a\u53f7\u7a2e\u5225"
PRIORITY_DISTRICTS = {"\u86f8\u5cf6", "\u6b63\u9662", "\u98ef\u7530", "\u4e0a\u6238", "\u76f4", "\u5b9d\u7acb"}


def meridional_arc(phi: float, a: float, e2: float) -> float:
    e4 = e2 * e2
    e6 = e4 * e2
    return a * (
        (1 - e2 / 4 - 3 * e4 / 64 - 5 * e6 / 256) * phi
        - (3 * e2 / 8 + 3 * e4 / 32 + 45 * e6 / 1024) * math.sin(2 * phi)
        + (15 * e4 / 256 + 45 * e6 / 1024) * math.sin(4 * phi)
        - (35 * e6 / 3072) * math.sin(6 * phi)
    )


def inverse_jprcs_vii(easting: float, northing: float) -> tuple[float, float]:
    # EPSG:6675 = JGD2011 / Japan Plane Rectangular CS VII.
    # Origin: 36N, 137d10E, scale 0.9999, GRS80 ellipsoid.
    a = 6378137.0
    inv_f = 298.257222101
    f = 1 / inv_f
    e2 = 2 * f - f * f
    ep2 = e2 / (1 - e2)
    k0 = 0.9999
    lat0 = math.radians(36.0)
    lon0 = math.radians(137.0 + 10.0 / 60.0)

    m0 = meridional_arc(lat0, a, e2)
    m = m0 + northing / k0
    mu = m / (a * (1 - e2 / 4 - 3 * e2 * e2 / 64 - 5 * e2**3 / 256))

    e1 = (1 - math.sqrt(1 - e2)) / (1 + math.sqrt(1 - e2))
    phi1 = (
        mu
        + (3 * e1 / 2 - 27 * e1**3 / 32) * math.sin(2 * mu)
        + (21 * e1 * e1 / 16 - 55 * e1**4 / 32) * math.sin(4 * mu)
        + (151 * e1**3 / 96) * math.sin(6 * mu)
        + (1097 * e1**4 / 512) * math.sin(8 * mu)
    )

    sin1 = math.sin(phi1)
    cos1 = math.cos(phi1)
    tan1 = math.tan(phi1)
    n1 = a / math.sqrt(1 - e2 * sin1 * sin1)
    r1 = a * (1 - e2) / (1 - e2 * sin1 * sin1) ** 1.5
    t1 = tan1 * tan1
    c1 = ep2 * cos1 * cos1
    d = easting / (n1 * k0)

    lat = phi1 - (n1 * tan1 / r1) * (
        d * d / 2
        - (5 + 3 * t1 + 10 * c1 - 4 * c1 * c1 - 9 * ep2) * d**4 / 24
        + (61 + 90 * t1 + 298 * c1 + 45 * t1 * t1 - 252 * ep2 - 3 * c1 * c1) * d**6 / 720
    )
    lon = lon0 + (
        d
        - (1 + 2 * t1 + c1) * d**3 / 6
        + (5 - 2 * c1 + 28 * t1 - 3 * c1 * c1 + 8 * ep2 + 24 * t1 * t1) * d**5 / 120
    ) / cos1
    return math.degrees(lon), math.degrees(lat)


def parse_gpkg_geometry(blob: bytes) -> list[list[float]]:
    flags = blob[3]
    endian = "<" if flags & 1 else ">"
    envelope_code = (flags >> 1) & 0b111
    offset = 8 + {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}.get(envelope_code, 0)
    wkb = blob[offset:]
    wkb_endian = "<" if wkb[0] == 1 else ">"
    geom_type = struct.unpack(wkb_endian + "I", wkb[1:5])[0]
    if geom_type != 2:
        raise ValueError(f"Unsupported WKB geometry type: {geom_type}")
    count = struct.unpack(wkb_endian + "I", wkb[5:9])[0]
    coords = []
    pos = 9
    for _ in range(count):
        x, y = struct.unpack(wkb_endian + "dd", wkb[pos : pos + 16])
        pos += 16
        coords.append(list(inverse_jprcs_vii(x, y)))
    return coords


def read_qgis_features() -> list[dict]:
    con = sqlite3.connect(GPKG_PATH)
    cur = con.cursor()
    table, geom_col = cur.execute("select table_name,column_name from gpkg_geometry_columns").fetchone()
    cols = [row[1] for row in cur.execute(f'pragma table_info("{table}")').fetchall()]
    rows = cur.execute(f'select * from "{table}"').fetchall()
    con.close()

    features = []
    for row in rows:
        record = dict(zip(cols, row))
        coords = parse_gpkg_geometry(record[geom_col])
        props = {k: v for k, v in record.items() if k != geom_col}
        features.append({"properties": props, "geometry": {"type": "LineString", "coordinates": coords}})
    return features


def read_construction_rows() -> dict[str, dict]:
    wb = load_workbook(QGIS_EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        item_id = str(record.get(SRC_ID) or "").strip()
        if item_id:
            rows[item_id] = record
    return rows


def load_base_geojson() -> dict:
    return json.loads(BASE_GEOJSON_PATH.read_text(encoding="utf-8"))


def make_feature(qgis_feature: dict, excel_row: dict | None) -> dict:
    props = qgis_feature["properties"]
    item_id = str(props.get(SRC_ID) or "").strip()
    raw_district = props.get(SRC_DISTRICT) or ""
    district = raw_district if raw_district in PRIORITY_DISTRICTS else (excel_row or {}).get(EXCEL_DISTRICT) or raw_district or "正院"
    record_id = f"R-REAL-{item_id}" if item_id else f"R-REAL-F{props.get('fid')}"
    road_name = props.get(SRC_ROAD_NAME) or (excel_row or {}).get(EXCEL_ROAD_NAME) or ""
    assessment = props.get(SRC_ASSESSMENT) or (excel_row or {}).get(EXCEL_ASSESSMENT) or ""
    segment = props.get(SRC_SEGMENT)
    title = (excel_row or {}).get(EXCEL_ITEM_NAME) or road_name or f"Real road {record_id}"
    return {
        "type": "Feature",
        "properties": {
            KEY_ID: record_id,
            KEY_DISTRICT: district,
            KEY_SOURCE: "QGIS real road geometry",
            KEY_EXTRACT: "real road replacement",
            "QGIS_ID": item_id,
            "QGIS_fid": props.get("fid"),
            "QGIS_segment": segment,
            COL_ITEM_STATUS: "\u5019\u88dc",
            APP_REAL_ID: item_id,
            "番号種別": (excel_row or {}).get(EXCEL_NUMBER_TYPE) or "",
            "道路名称": road_name,
            "査定番号": str(assessment),
            "箇所名": (excel_row or {}).get(EXCEL_PLACE) or "",
            "復旧延長_m": (excel_row or {}).get(EXCEL_LENGTH) or "",
            "幅員_m": (excel_row or {}).get(EXCEL_WIDTH) or "",
            "工事内容": (excel_row or {}).get(EXCEL_CONTENT) or "",
            "工事区分": props.get(SRC_WORK) or (excel_row or {}).get(EXCEL_WORK) or "",
            "工事名": title,
            "実データ": "true",
        },
        "geometry": qgis_feature["geometry"],
    }


def normalize_work_type(value: str) -> str:
    value = str(value or "").strip()
    if "橋梁" in value:
        return "橋梁工事"
    if "舗装" in value:
        return "舗装工事"
    if "道路" in value:
        return "道路復旧工事"
    if "下水" in value:
        return "下水道工事"
    if "排水" in value:
        return "排水工事"
    return value or "道路復旧工事"


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def create_excel(qgis_features: list[dict], construction_rows: dict[str, dict]) -> None:
    wb = load_workbook(BASE_EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    needed = [
        "施工内容",
        "施工延長",
        "道路種別",
        "道路幅員_前",
        "道路幅員_後",
        "舗装構成",
        "歩道",
        "区画線",
        COL_ITEM_STATUS,
        APP_REAL_ID,
        "番号種別",
        "査定番号",
        "箇所名",
        "復旧延長_m",
        "幅員_m",
        "工事区分",
        "道路名称",
        "QGIS_ID",
    ]
    for header in needed:
        if header not in headers:
            ws.cell(row=1, column=len(headers) + 1, value=header)
            copy_cell_style(ws.cell(row=1, column=1), ws.cell(row=1, column=len(headers) + 1))
            headers.append(header)

    existing_ids = {str(row[0].value).strip() for row in ws.iter_rows(min_row=2, max_col=1)}
    unique_item_ids = []
    fallback_features = []
    for feature in qgis_features:
        item_id = str(feature["properties"].get(SRC_ID) or "").strip()
        if item_id and item_id not in unique_item_ids:
            unique_item_ids.append(item_id)
        elif not item_id:
            fallback_features.append(feature)

    row_sources = [("id", item_id, None) for item_id in unique_item_ids]
    row_sources.extend(("fid", str(feature["properties"].get("fid")), feature) for feature in fallback_features)

    for source_kind, item_id, fallback_feature in row_sources:
        record_id = f"R-REAL-{item_id}" if source_kind == "id" else f"R-REAL-F{item_id}"
        if record_id in existing_ids:
            continue
        source = construction_rows.get(item_id, {}) if source_kind == "id" else {}
        fallback_props = (fallback_feature or {}).get("properties", {})
        work_type = normalize_work_type(source.get(EXCEL_WORK) or fallback_props.get(SRC_WORK))
        road_name = source.get(EXCEL_ROAD_NAME) or fallback_props.get(SRC_ROAD_NAME) or source.get(EXCEL_ITEM_NAME) or ""
        title = source.get(EXCEL_ITEM_NAME) or road_name or f"正院 real road {item_id}"
        length = source.get(EXCEL_LENGTH) or ""
        width = source.get(EXCEL_WIDTH) or ""
        before_width = f"{width}m" if width not in ("", None) else ""
        values = {
            KEY_ID: record_id,
            COL_TITLE: title,
            COL_WORK_TYPE: "",
            COL_RESTRICTION_TYPE: "",
            COL_START: datetime(2026, 7, 15),
            COL_END: datetime(2026, 12, 31),
            COL_CONTRACTOR: "未設定",
            COL_PROGRESS: 0,
            COL_NOTE: "",
            "施工内容": source.get(EXCEL_CONTENT) or "道路復旧",
            "施工延長": length,
            "道路種別": "市道",
            "道路幅員_前": before_width,
            "道路幅員_後": before_width,
            "舗装構成": "As舗装",
            "歩道": "既設利用",
            "区画線": "施工予定",
            COL_ITEM_STATUS: "\u5019\u88dc",
            APP_REAL_ID: item_id if source_kind == "id" else "",
            "番号種別": source.get(EXCEL_NUMBER_TYPE) or "",
            "査定番号": source.get(EXCEL_ASSESSMENT) or fallback_props.get(SRC_ASSESSMENT) or "",
            "箇所名": source.get(EXCEL_PLACE) or "",
            "復旧延長_m": length,
            "幅員_m": width,
            "工事区分": source.get(EXCEL_WORK) or fallback_props.get(SRC_WORK) or "",
            "道路名称": road_name,
            "QGIS_ID": item_id if source_kind == "id" else "",
        }
        row_idx = ws.max_row + 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=values.get(header, ""))
            copy_cell_style(ws.cell(row=2, column=col_idx), ws.cell(row=row_idx, column=col_idx))

    wb.save(OUT_EXCEL_PATH)


def main() -> None:
    construction_rows = read_construction_rows()
    qgis_features = [
        feature
        for feature in read_qgis_features()
        if feature["properties"].get(SRC_ID) or feature["properties"].get(SRC_ROAD_NAME)
    ]

    base_geojson = load_base_geojson()
    real_features = [
        make_feature(feature, construction_rows.get(str(feature["properties"].get(SRC_ID) or "").strip()))
        for feature in qgis_features
    ]
    replacement_districts = {
        feature.get("properties", {}).get(KEY_DISTRICT)
        for feature in real_features
        if feature.get("properties", {}).get(KEY_DISTRICT) in PRIORITY_DISTRICTS
    }
    kept = [
        feature
        for feature in base_geojson.get("features", [])
        if feature.get("properties", {}).get(KEY_DISTRICT) not in replacement_districts
    ]
    base_geojson["features"] = kept + real_features
    OUT_GEOJSON_PATH.write_text(json.dumps(base_geojson, ensure_ascii=False, indent=2), encoding="utf-8")
    create_excel(qgis_features, construction_rows)

    lons = [coord[0] for feature in real_features for coord in feature["geometry"]["coordinates"]]
    lats = [coord[1] for feature in real_features for coord in feature["geometry"]["coordinates"]]
    print(f"real features: {len(real_features)}")
    print(f"base kept features: {len(kept)}")
    print(f"output features: {len(base_geojson['features'])}")
    print(f"bbox lon/lat: {min(lons):.6f},{min(lats):.6f},{max(lons):.6f},{max(lats):.6f}")
    print(OUT_GEOJSON_PATH)
    print(OUT_EXCEL_PATH)


if __name__ == "__main__":
    main()
