from __future__ import annotations

import csv
import json
import collections
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / ".codex-tmp" / "master_excel"
SOURCE_XLSX = ROOT / "data" / "excel" / "qgis_construction_list.xlsx"
GEOJSON_PATH = ROOT / "data" / "geojson" / "suzu_with_real_roads.geojson"

J = {
    "item": "\u500b\u5225\u9805\u76ee\u540d",
    "road": "\u500b\u5225\u9053\u8def\u540d",
    "bridge": "\u6a4b\u6881\u540d",
    "district": "\u5730\u533a",
    "work": "\u5de5\u4e8b\u533a\u5206",
    "num_type": "\u756a\u53f7\u7a2e\u5225",
    "assessment": "\u67fb\u5b9a\u756a\u53f7",
    "place": "\u7b87\u6240\u540d",
    "length": "\u5fa9\u65e7\u5ef6\u9577_m",
    "width": "\u5e45\u54e1_m",
    "subno": "\u5143\u884c\u5185\u9023\u756a",
    "subtotal": "\u5143\u884c\u5185\u9805\u76ee\u6570",
    "content": "\u5de5\u4e8b\u5185\u5bb9",
    "start": "\u59cb\u70b9",
    "end": "\u7d42\u70b9",
    "feature_id": "\u898f\u5236ID",
    "priority": "\u91cd\u70b9\u5730\u533a",
    "road_name": "\u9053\u8def\u540d\u79f0",
}

MASTER_HEADERS = [
    "ID",
    J["item"],
    J["road"],
    J["bridge"],
    J["district"],
    J["work"],
    J["num_type"],
    J["assessment"],
    J["place"],
    J["length"],
    J["width"],
    J["subno"],
    J["subtotal"],
    J["content"],
    J["start"],
    J["end"],
    "geometry_status",
    "geometry_feature_count",
    "app_feature_ids",
    "qgis_fids",
    "qgis_segments",
    "geometry_point_count",
    "start_lon",
    "start_lat",
    "end_lon",
    "end_lat",
    "source_workbook",
]

APP_HEADERS = [
    "app_feature_id",
    "ID",
    "qgis_fid",
    "qgis_segment",
    J["road_name"],
    J["assessment"],
    J["district"],
    J["work"],
    "geometry_point_count",
    "start_lon",
    "start_lat",
    "end_lon",
    "end_lat",
]


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def qid(value):
    if value in (None, ""):
        return ""
    try:
        as_float = float(value)
        if as_float.is_integer():
            return str(int(as_float))
    except Exception:
        pass
    return clean(value)


def write_csv(path: Path, headers: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def source_rows() -> list[dict]:
    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    index = {header: i for i, header in enumerate(headers) if header}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {header: row[i] if i < len(row) else None for header, i in index.items()}
        if record.get("ID"):
            rows.append(record)
    return rows


def geometry_rows() -> tuple[dict[str, list[dict]], list[dict]]:
    data = json.loads(GEOJSON_PATH.read_text(encoding="utf-8"))
    grouped = collections.defaultdict(list)
    app_rows = []
    for feature in data.get("features", []):
        props = feature.get("properties", {})
        if props.get("QGIS_fid") in ("", None):
            continue
        qgis_id = qid(props.get("QGIS_ID"))
        coords = feature.get("geometry", {}).get("coordinates") or []
        row = {
            "app_feature_id": clean(props.get(J["feature_id"])),
            "ID": qgis_id,
            "qgis_fid": qid(props.get("QGIS_fid")),
            "qgis_segment": qid(props.get("QGIS_segment")),
            J["road_name"]: clean(props.get(J["road_name"])),
            J["assessment"]: clean(props.get(J["assessment"])),
            J["district"]: clean(props.get(J["priority"])),
            J["work"]: clean(props.get(J["work"])),
            "geometry_point_count": len(coords),
            "start_lon": coords[0][0] if coords else "",
            "start_lat": coords[0][1] if coords else "",
            "end_lon": coords[-1][0] if coords else "",
            "end_lat": coords[-1][1] if coords else "",
        }
        app_rows.append(row)
        if qgis_id:
            grouped[qgis_id].append(row)
    return grouped, app_rows


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = source_rows()
    grouped_geometry, app_rows = geometry_rows()

    master = []
    for record in rows:
        item_id = qid(record.get("ID"))
        geoms = grouped_geometry.get(item_id, [])
        first = geoms[0] if geoms else {}
        last = geoms[-1] if geoms else {}
        master.append(
            {
                "ID": item_id,
                J["item"]: clean(record.get(J["item"])),
                J["road"]: clean(record.get(J["road"])),
                J["bridge"]: clean(record.get(J["bridge"])),
                J["district"]: clean(record.get(J["district"])),
                J["work"]: clean(record.get(J["work"])),
                J["num_type"]: clean(record.get(J["num_type"])),
                J["assessment"]: clean(record.get(J["assessment"])),
                J["place"]: clean(record.get(J["place"])),
                J["length"]: record.get(J["length"]) if record.get(J["length"]) is not None else "",
                J["width"]: record.get(J["width"]) if record.get(J["width"]) is not None else "",
                J["subno"]: record.get(J["subno"]) if record.get(J["subno"]) is not None else "",
                J["subtotal"]: record.get(J["subtotal"]) if record.get(J["subtotal"]) is not None else "",
                J["content"]: clean(record.get(J["content"])),
                J["start"]: clean(record.get(J["start"])),
                J["end"]: clean(record.get(J["end"])),
                "geometry_status": "mapped" if geoms else "not_mapped_yet",
                "geometry_feature_count": len(geoms),
                "app_feature_ids": "; ".join(g["app_feature_id"] for g in geoms if g["app_feature_id"]),
                "qgis_fids": "; ".join(g["qgis_fid"] for g in geoms if g["qgis_fid"]),
                "qgis_segments": "; ".join(g["qgis_segment"] for g in geoms if g["qgis_segment"]),
                "geometry_point_count": sum(int(g["geometry_point_count"]) for g in geoms),
                "start_lon": first.get("start_lon", ""),
                "start_lat": first.get("start_lat", ""),
                "end_lon": last.get("end_lon", ""),
                "end_lat": last.get("end_lat", ""),
                "source_workbook": "qgis_construction_list.xlsx",
            }
        )

    summary = [
        {"metric": "source_rows", "value": len(master)},
        {"metric": "mapped_rows", "value": sum(1 for row in master if row["geometry_status"] == "mapped")},
        {"metric": "not_mapped_yet_rows", "value": sum(1 for row in master if row["geometry_status"] != "mapped")},
        {"metric": "qgis_geometry_features", "value": len(app_rows)},
    ]
    for district, count in collections.Counter(row[J["district"]] for row in master).items():
        summary.append({"metric": f"source_rows_district_{district}", "value": count})
    for district, count in collections.Counter(row[J["district"]] for row in master if row["geometry_status"] == "mapped").items():
        summary.append({"metric": f"mapped_rows_district_{district}", "value": count})

    write_csv(OUT_DIR / "master_rows.csv", MASTER_HEADERS, master)
    write_csv(OUT_DIR / "app_feature_map.csv", APP_HEADERS, app_rows)
    write_csv(OUT_DIR / "summary.csv", ["metric", "value"], summary)
    print(f"master rows: {len(master)}")
    print(f"mapped rows: {sum(1 for row in master if row['geometry_status'] == 'mapped')}")
    print(f"geometry features: {len(app_rows)}")


if __name__ == "__main__":
    main()
